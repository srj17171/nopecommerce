import openpyxl

path = "C:\\Users\\suraj\\Desktop\\automation project\\new project nope commerce\\testdata\\data2.xlsx"

workbook = openpyxl.load_workbook(path)
sheet1 = workbook['Sheet1']
print(sheet1.cell(row = 1, column = 2).value)
sheet1.cell(row = 1, column = 2).value = 9873764

sheet1.cell(row = 2, column = 1).value = "vijay"
sheet1.cell(row = 3, column = 1).value = 'sonam'
sheet1.cell(row = 2, column = 2).value = 54698754
sheet1.cell(row = 3, column = 2).value = 4578512
workbook.save(path)
