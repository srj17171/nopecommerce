import openpyxl
# --- for creating a workbook
workbook = openpyxl.load_workbook("data.xlsx")
# ---- for selecting a sheet
sheet = workbook["Sheet1"]
sheet2 = workbook.active
# ----for geeting a value
print(sheet2.cell(row=1, column=1).value)
# for changing a value in the sheet
sheet2.cell(row = 1, column = 1).value = "name of user"
workbook.save("data.xlsx")
# print(sheet2.cell(row=1, column=1).value)

# detail = sheet2.cell(row = 1, column = 1)
# workbook.save("data.xlsx")
# print(detail.value)
# for i in range(2, sheet2.max_row + 1):
#     for j in range(3, 5):
#         detail = sheet2.cell(row=i, column = j)
#         print(detail.value)


def rwo_num(file,Sheet_name):
    book = openpyxl.load_workbook(file)
    sheet = book[Sheet_name]
    return sheet.max_row
# print(rwo_num("data.xlsx","Sheet1"))
#
def read_data(file, Sheet_name, rownum, columnnum):
    book = openpyxl.load_workbook(file)
    sheet = book[Sheet_name]
    return sheet.cell(row = rownum, column = columnnum).value
    #
    # print(read_data("data.xlsx", "Sheet1", 1, 1))

def write_data(file, Sheet_name, rownum, columnnum, data):
    book = openpyxl.load_workbook(file)
    sheet = book[Sheet_name]
    sheet.cell(row = rownum, column = columnnum).value = data
    book.save(file)
#
# write_data("data.xlsx", "Sheet1", 1, 1, "name")