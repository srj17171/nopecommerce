import openpyxl


def row_num(file, Sheet_name):
    book = openpyxl.load_workbook(file)
    sheet = book[Sheet_name]
    return sheet.max_row


def colomn_num(file, sheet_name):
    book = openpyxl.load_workbook(file)
    sheet = book[sheet_name]
    return sheet.max_coloumn


def read_data(file, Sheet_name, rownum, columnnum):
    book = openpyxl.load_workbook(file)
    sheet = book[Sheet_name]
    return sheet.cell(row = rownum, column = columnnum).value


def write_data(file, Sheet_name, rownum, columnnum, data):
    book = openpyxl.load_workbook(file)
    sheet = book[Sheet_name]
    sheet.cell(row = rownum, column = columnnum).value = data
    book.save(file)
